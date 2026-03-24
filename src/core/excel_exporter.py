"""
excel_exporter.py
=================
Exports DBC data (messages + signals) to XLSX (Excel) or CSV format.
"""

from __future__ import annotations

import csv
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Shared header style colours
_HDR_BG = "1F3864"  # dark-blue
_HDR_FG = "FFFFFF"  # white


def _apply_header(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, color=_HDR_FG)
    cell.fill = PatternFill(start_color=_HDR_BG, end_color=_HDR_BG, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws) -> None:
    """Widen every column to its longest value (capped at 60)."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 3, 60
        )


class ExcelExporter:
    """Static helpers for exporting :class:`~core.dbc_parser.MessageData` lists."""

    # ------------------------------------------------------------------
    # XLSX
    # ------------------------------------------------------------------

    @staticmethod
    def export_xlsx(messages: List, file_path: str) -> None:  # type: ignore[type-arg]
        wb = openpyxl.Workbook()

        # ---- Summary sheet -----------------------------------------------
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum["A1"] = "FreeDBC Editor & Previewer — Export"
        ws_sum["A1"].font = Font(bold=True, size=14, color="1F3864")
        ws_sum["A3"] = "Total Messages"
        ws_sum["B3"] = len(messages)
        ws_sum["A4"] = "Total Signals"
        ws_sum["B4"] = sum(len(m.signals) for m in messages)
        ws_sum["A3"].font = ws_sum["A4"].font = Font(bold=True)

        # ---- Messages sheet -----------------------------------------------
        ws_msg = wb.create_sheet("Messages")
        msg_hdrs = [
            "ID (hex)",
            "ID (dec)",
            "Name",
            "DLC (bytes)",
            "Signal Count",
            "Comment",
        ]
        for col, hdr in enumerate(msg_hdrs, 1):
            _apply_header(ws_msg.cell(row=1, column=col), hdr)
        ws_msg.row_dimensions[1].height = 22

        for row, msg in enumerate(messages, 2):
            ws_msg.cell(row=row, column=1, value=f"0x{msg.id:03X}")
            ws_msg.cell(row=row, column=2, value=msg.id)
            ws_msg.cell(row=row, column=3, value=msg.name)
            ws_msg.cell(row=row, column=4, value=msg.length)
            ws_msg.cell(row=row, column=5, value=len(msg.signals))
            ws_msg.cell(row=row, column=6, value=msg.comment or "")
        _autofit(ws_msg)

        # ---- Signals sheet -----------------------------------------------
        ws_sig = wb.create_sheet("Signals")
        sig_hdrs = [
            "Message Name",
            "Message ID (hex)",
            "Signal Name",
            "Start Bit",
            "Bit Length",
            "Byte Order",
            "Signed",
            "Factor",
            "Offset",
            "Min",
            "Max",
            "Unit",
            "Comment",
            "Value Descriptions",
        ]
        for col, hdr in enumerate(sig_hdrs, 1):
            _apply_header(ws_sig.cell(row=1, column=col), hdr)
        ws_sig.row_dimensions[1].height = 22

        sig_row = 2
        for msg in messages:
            for sig in msg.signals:
                val_desc = (
                    "; ".join(f"{k}={v}" for k, v in sorted(sig.choices.items()))
                    if sig.choices
                    else ""
                )
                data = [
                    msg.name,
                    f"0x{msg.id:03X}",
                    sig.name,
                    sig.start_bit,
                    sig.length,
                    (
                        "Little-Endian"
                        if sig.byte_order == "little_endian"
                        else "Big-Endian"
                    ),
                    "Yes" if sig.is_signed else "No",
                    sig.factor,
                    sig.offset,
                    sig.minimum if sig.minimum is not None else "",
                    sig.maximum if sig.maximum is not None else "",
                    sig.unit,
                    sig.comment or "",
                    val_desc,
                ]
                for col, value in enumerate(data, 1):
                    ws_sig.cell(row=sig_row, column=col, value=value)
                sig_row += 1
        _autofit(ws_sig)

        # ---- Per-message sheets ------------------------------------------
        for msg in messages:
            if not msg.signals:
                continue
            sheet_name = msg.name[:31]  # Excel name limit
            ws = wb.create_sheet(sheet_name)
            per_hdrs = [
                "Signal Name",
                "Start Bit",
                "Bit Length",
                "Byte Order",
                "Signed",
                "Factor",
                "Offset",
                "Min",
                "Max",
                "Unit",
                "Comment",
                "Value Descriptions",
            ]
            for col, hdr in enumerate(per_hdrs, 1):
                _apply_header(ws.cell(row=1, column=col), hdr)
            for row_idx, sig in enumerate(msg.signals, 2):
                val_desc = (
                    "; ".join(f"{k}={v}" for k, v in sorted(sig.choices.items()))
                    if sig.choices
                    else ""
                )
                row_data = [
                    sig.name,
                    sig.start_bit,
                    sig.length,
                    "LE" if sig.byte_order == "little_endian" else "BE",
                    "Yes" if sig.is_signed else "No",
                    sig.factor,
                    sig.offset,
                    sig.minimum if sig.minimum is not None else "",
                    sig.maximum if sig.maximum is not None else "",
                    sig.unit,
                    sig.comment or "",
                    val_desc,
                ]
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col, value=value)
            _autofit(ws)

        wb.save(file_path)

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    @staticmethod
    def export_csv(messages: List, file_path: str) -> None:  # type: ignore[type-arg]
        headers = [
            "Message Name",
            "Message ID",
            "Signal Name",
            "Start Bit",
            "Bit Length",
            "Byte Order",
            "Signed",
            "Factor",
            "Offset",
            "Min",
            "Max",
            "Unit",
            "Comment",
            "Value Descriptions",
        ]
        with open(file_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            for msg in messages:
                for sig in msg.signals:
                    val_desc = (
                        "; ".join(f"{k}={v}" for k, v in sorted(sig.choices.items()))
                        if sig.choices
                        else ""
                    )
                    writer.writerow(
                        [
                            msg.name,
                            f"0x{msg.id:03X}",
                            sig.name,
                            sig.start_bit,
                            sig.length,
                            (
                                "Little-Endian"
                                if sig.byte_order == "little_endian"
                                else "Big-Endian"
                            ),
                            "Yes" if sig.is_signed else "No",
                            sig.factor,
                            sig.offset,
                            sig.minimum if sig.minimum is not None else "",
                            sig.maximum if sig.maximum is not None else "",
                            sig.unit,
                            sig.comment or "",
                            val_desc,
                        ]
                    )
